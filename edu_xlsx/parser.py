from pathlib import Path
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import _STRING_COL_CACHE
from itertools import chain as itertools_chain
from json import dump

from .models import Class, Subject, Classroom, Teacher, Group, Lesson, Card
from .utils import build_ids, build_groups, get_short_first_chars, parse_classroom_names, parse_coordinate_ord

from typing import List, Dict, Tuple, Any, Union, Optional


ACCEPTABLE_TIMETABLE_NUMBERS: List[str] = [
    "1",
    "2"
]


class XLSXParser:
    def __init__(self, xlsx_filepath: Union[str, Path], timetable_number: Union[int, str]) -> None:
        if not isinstance(xlsx_filepath, Path):
            xlsx_filepath = Path(xlsx_filepath)

        if not isinstance(timetable_number, str):
            timetable_number = str(timetable_number)

        self.xlsx_filepath: Path = xlsx_filepath
        self.timetable_number: str = timetable_number

        if not self.xlsx_filepath.exists() or not self.xlsx_filepath.is_file():
            raise ValueError(
                "Input .XLSX `{xlsx_filepath}` must exists and be a file!".format(
                    xlsx_filepath = xlsx_filepath
                )
            )

        if timetable_number not in ACCEPTABLE_TIMETABLE_NUMBERS:
            raise ValueError(
                "TimeTable number `{timetable_number}` must be {correct_values}!".format(
                    timetable_number = timetable_number,
                    correct_values = " or ".join(ACCEPTABLE_TIMETABLE_NUMBERS)
                )
            )

        if timetable_number == ACCEPTABLE_TIMETABLE_NUMBERS[0]:
            from .datas_1 import periods, daysdefs

        elif timetable_number == ACCEPTABLE_TIMETABLE_NUMBERS[1]:
            from .datas_2 import periods, daysdefs

        self._periods: List[dict] = periods
        self._periods_len: int = len(self._periods)
        self._daysdefs: List[dict] = daysdefs
        self._daysdefs_len: int = len(self._daysdefs)

        self._excel: Workbook = load_workbook(
            filename = xlsx_filepath
        )

        self._sheet: Worksheet = self._excel[self._excel.sheetnames[0]]

        self._sheet_max_row: int = self._sheet.max_row
        self._sheet_max_column: int = self._sheet.max_column
        self._sheet_max_column_letters: str = _STRING_COL_CACHE[self._sheet_max_column]

    def parse(self) -> None:
        i: int
        id: str

        self._classes: List[Class] = []
        self._classes_from_id: Dict[str, Class] = {}
        self._classes_from_name: Dict[str, Class] = {}

        i = 1

        for _classes in self._sheet["A5:A{}".format(self._sheet_max_row)]:
            for class_cell in _classes:
                if not class_cell.value:
                    continue

                class_name: str = class_cell.value.strip()

                id = "*{}".format(i)

                class_: Class = Class(
                    id = id,
                    name = class_name,
                    short = class_name
                )

                self._classes.append(class_)
                self._classes_from_id[id] = class_
                self._classes_from_name[class_name] = class_

                i += 1

        self._groups_from_class_id: Dict[str, Tuple[Group, Group, Group]] = {}
        self._groups: list[Group] = []

        for group_ in build_ids(
            items = build_groups(
                classids = self._classes_from_id.keys()
            )
        ):
            group: Group = Group(
                id = group_["id"],
                name = group_["name"],
                classid = group_["classid"],
                entireclass = group_["entireclass"],
                ascttdivision = (
                    ""
                    if group_["divisiontag"] == "0"
                    else
                    group_["divisiontag"]
                ),
                divisionid = "*{0}:{1}".format(
                    i,
                    (
                        ""
                        if group_["entireclass"]
                        else
                        "1"
                    )
                )
            )

            self._groups.append(group)

            if group.classid in self._groups_from_class_id:
                self._groups_from_class_id[group.classid].append(group)
            else:
                self._groups_from_class_id[group.classid] = [group]

        sheet_table_datas: Dict[str, List[Any]] = {
            "subjects": [],
            "classrooms": [],
            "teachers": [],
            "lessons": []
        }

        for i in range(0, len(self._classes)):
            i_x4: int = i * 4

            sheet_table_datas["subjects"].append(self._sheet["B{0}:{1}{0}".format(5 + i_x4, self._sheet_max_column_letters)])
            sheet_table_datas["classrooms"].append(self._sheet["B{0}:{1}{0}".format(7 + i_x4, self._sheet_max_column_letters)])
            sheet_table_datas["teachers"].append(self._sheet["B{0}:{1}{0}".format(8 + i_x4, self._sheet_max_column_letters)])
            sheet_table_datas["lessons"].append(self._sheet["B{0}:{1}{2}".format(5 + i_x4, self._sheet_max_column_letters, 5 + i_x4 + 3)])

        self._subjects: List[Subject] = []
        self._subjects_from_id: Dict[str, Subject] = {}
        self._subjects_from_name: Dict[str, Subject] = {}

        i = 1

        for _subjects in sheet_table_datas["subjects"]:
            for __subjects in _subjects:
                for subject_cell in __subjects:
                    subject_name: Union[str, None] = subject_cell.value

                    if not subject_name:
                        continue

                    subject_name = subject_name.strip()

                    if subject_name in self._subjects_from_name:
                        continue

                    id = "*{}".format(i)

                    subject: Subject = Subject(
                        id = id,
                        name = subject_name,
                        short = get_short_first_chars(
                            string = subject_name
                        )
                    )

                    self._subjects.append(subject)
                    self._subjects_from_id[id] = subject
                    self._subjects_from_name[subject_name] = subject

                    i += 1

        self._classrooms: List[Classroom] = []
        self._classrooms_from_id: Dict[str, Classroom] = {}
        self._classrooms_from_name: Dict[str, Classroom] = {}

        i = 1

        for _classrooms in sheet_table_datas["classrooms"]:
            for __classrooms in _classrooms:
                for classroom_cell in __classrooms:
                    classroom_names_str: Union[str, int, None] = classroom_cell.value

                    if not classroom_names_str:
                        continue

                    classroom_names_str = str(classroom_names_str)

                    for classroom_name in parse_classroom_names(
                        classroom_names_str = classroom_names_str,
                        classrooms_from_name = self._classrooms_from_name
                    ):
                        id = "*{}".format(i)

                        classroom: Classroom = Classroom(
                            id = id,
                            name = classroom_name,
                            short = classroom_name
                        )

                        self._classrooms.append(classroom)
                        self._classrooms_from_id[id] = classroom
                        self._classrooms_from_name[classroom_name] = classroom

                        i += 1

        self._teachers: List[Teacher] = []
        self._teachers_from_id: Dict[str, Teacher] = {}
        self._teachers_from_name: Dict[str, Teacher] = {}

        i = 1

        for _teachers in sheet_table_datas["teachers"]:
            for __teachers in _teachers:
                for teachers_cell in __teachers:
                    _teachers_names: Union[str, None] = teachers_cell.value

                    if not _teachers_names:
                        continue

                    for teacher_name in _teachers_names.split(","):
                        teacher_name = teacher_name.strip()

                        if teacher_name in self._teachers_from_name:
                            continue

                        id = "*{}".format(i)

                        teacher: Teacher = Teacher(
                            id = id,
                            firstname = "",
                            lastname = teacher_name,
                            short = "" # TODO: every first letter
                            # TODO: maybe use `gender`
                        )

                        self._teachers.append(teacher)
                        self._teachers_from_id[id] = teacher
                        self._teachers_from_name[teacher_name] = teacher

                        i += 1

        self._lessons: List[Lesson] = []
        self._lessons_from_id: Dict[str, Lesson] = {}

        self._cards: List[Card] = []

        i: int = 1

        for sheet_lessons_rows in sheet_table_datas["lessons"]:
            for subject_cell, classes_cell, classrooms_cell, teachers_cell in zip(sheet_lessons_rows[0], sheet_lessons_rows[1], sheet_lessons_rows[2], sheet_lessons_rows[3]):
                if not subject_cell.value:
                    continue

                coordinate_ord: int = parse_coordinate_ord(
                    coordinate = subject_cell.coordinate
                )

                period: int = (coordinate_ord - 65) % self._periods_len or self._periods_len
                daysdef_id_num: int = (coordinate_ord - 65 + self._periods_len - 1) // self._periods_len

                id = "*{}".format(i)

                classid: str = self._classes_from_name[classes_cell.value.split(",")[0].strip()].id
                classroom_names_str: Union[str, int, None] = classrooms_cell.value

                if not classroom_names_str:
                    classroom_names_str = ""

                classroom_names_str = str(classroom_names_str)

                classroomids: list[str] = []

                for classroom_name in parse_classroom_names(
                    classroom_names_str = classroom_names_str,
                    classrooms_from_name = self._classrooms_from_name
                ):
                    classroomids.append(self._classrooms_from_name[classroom_name].id)

                teacher_ids: list[str] = (
                    [
                        self._teachers_from_name[teacher_name.strip()].id
                        for teacher_name in teachers_cell.value.split(",")
                    ]
                    if teachers_cell.value
                    else
                    [""]
                )

                lesson: Lesson = Lesson(
                    id = id,
                    subjectid = self._subjects_from_name[subject_cell.value.strip()].id,
                    teacherids = teacher_ids,
                    groupids = [
                        group.id
                        for group in (
                            self._groups_from_class_id[classid][1:]
                            if len(teacher_ids) > 1
                            else
                            self._groups_from_class_id[classid][:1]
                        )
                    ],
                    classids = [classid],
                    classroomidss = [classroomids],
                    termsdefid = "*1",
                    weeksdefid = "*1",
                    daysdefid = "*{}".format(daysdef_id_num)
                )

                self._lessons.append(lesson)
                self._lessons_from_id[id] = lesson

                card: Card = Card(
                    id = id,
                    lessonid = lesson.id,
                    period = str(period),
                    days = self._daysdefs[daysdef_id_num - 1]["days"],
                    weeks = "1",
                    classroomids = lesson.classroomidss[0]
                )

                self._cards.append(card)

                i += 1

    def save(self, json_filepath: Union[str, Path], json_indent: Optional[int]=None) -> None:
        if not isinstance(json_filepath, Path):
            json_filepath = Path(json_filepath)

        if isinstance(json_indent, int) and json_indent <= 0:
            json_indent = None

        timetable_dict: dict = {
            "r": {
                "rights": {
                    "subjects": True,
                    "classes": True,
                    "teachers": True,
                    "classrooms": True,
                    "students": True,
                    "igroups": True,
                    "classroomsupervision": True,
                    "teachers_summary": True,
                    "classes_summary": True,
                    "classrooms_summary": True,
                    "igroups_summary": True
                },
                "dbiAccessorRes": {
                    "type": "ttuidocdbi",
                    "dbid": "",
                    "tables": [
                        {"id":"globals","def":{"id":"globals","name":"Расписание","item_name":"Расписание","icon":"/timetable/pics/app/office/school_32.png"},"cdefs":[{"id":"name","type":"string","name":"Название учреждения"},{"id":"settings","type":"subobject","name":"settings","subcolumns":[{"id":"m_nZlozitostGener","type":"intcombo","name":"m_nZlozitostGener"},{"id":"m_bAllowZlavnenie","type":"checkbox","name":"m_bAllowZlavnenie"},{"id":"m_bGenerDraft","type":"checkbox","name":"m_bGenerDraft"},{"id":"m_nCoGenerovat","type":"intcombo","name":"m_nCoGenerovat"},{"id":"m_nSchoolType","type":"intcombo","name":"Тип школы"},{"id":"m_nGapsCounting","type":"intcombo","name":"Способ подсчета окон"},{"id":"name_format","type":"combo","name":"Формат имени"},{"id":"m_strPrintHeaderText","type":"string","name":"Заголовок"},{"id":"m_strDateBellowTimeTable","type":"string","name":"m_strDateBellowTimeTable"},{"id":"m_bPrintDozory","type":"checkbox","name":"Печатать дежурства в индивидуальных расписаниях"},{"id":"m_bPrintDozoryVSuhrnnych","type":"checkbox","name":"Печатать дежурства в итоговых расписаниях"},{"id":"m_bPrintDozoryColor","type":"checkbox","name":"Печатать надсмотр над кабинетами в цвете"},{"id":"m_bPrintSinglesSpolu","type":"checkbox","name":"Печатать одиночные последоват. уроки как 1 урок"},{"id":"m_bPrintDoublesAsSingles","type":"checkbox","name":"Печатать спаренные уроки как одиночные"},{"id":"m_nTimeFormat","type":"intcombo","name":"Формат времени"},{"id":"m_nPrvyDen","type":"intcombo","name":"Выходные дни"},{"id":"m_bPrintDayAsNumber","type":"checkbox","name":"Показывать номер дня вместо дня недели (т.е. День 1 вместо Понед.)"},{"id":"m_DozoryKriteria","type":"subobject","name":"m_DozoryKriteria","subcolumns":[{"id":"ucitel_limit_pocet","type":"int","name":"Большее/меньшее число надсмотров чем определено в ограничениях учителей"},{"id":"ucitel_limit_minuty","type":"int","name":"Больше/меньше минут надсмотров чем определено в ограничениях учителей"},{"id":"ucitel_rovnomerne_pocet","type":"int","name":"Такое же число для всех учителей"},{"id":"ucitel_rovnomerne_minuty","type":"int","name":"Такое же число минут для всех учителей"},{"id":"ucitel_uci_predalebopo","type":"int","name":"Учитель имет урок до ИЛИ после"},{"id":"ucitel_uci_predajpo","type":"int","name":"Учитель имет урок до И после"},{"id":"ucitel_uci_vsusednej_predalebopo","type":"int","name":"В соседнем кабинете до ИЛИ после"},{"id":"ucitel_uci_vsusednej_predajpo","type":"int","name":"В соседнем кабинете до И после"},{"id":"ucitel_cezdvojhodinovku","type":"int","name":"Во время двойного (или более длинного) урока"},{"id":"ucitel_prazdnyden","type":"int","name":"В день без уроков"},{"id":"ucitel_predvyucovanim","type":"int","name":"Перед первым уроком учителя"},{"id":"ucitel_povyucovani","type":"int","name":"После последнего урока учителя"},{"id":"ucitel_vjedenden","type":"int","name":"Все в один день"},{"id":"ucitel_zasebou","type":"int","name":"Непрерывно"}]},{"id":"draft_options","type":"subdata","name":"draft_options","subcolumns":[{"id":"active","type":"checkbox","name":"active"},{"id":"relax","type":"checkbox","name":"relax"}]}]},{"id":"year","type":"schoolyear","name":"Учебный год"},{"id":"reg_name","type":"string","name":"reg_name"}],"data_rows":[{"id":"1","name":"","settings":{"m_nZlozitostGener":1,"m_bAllowZlavnenie":True,"m_bGenerDraft":False,"m_nCoGenerovat":0,"m_nSchoolType":0,"m_nGapsCounting":0,"name_format":"FSL","m_strPrintHeaderText":"","m_strDateBellowTimeTable":"","m_bPrintDozory":True,"m_bPrintDozoryVSuhrnnych":False,"m_bPrintDozoryColor":False,"m_bPrintSinglesSpolu":False,"m_bPrintDoublesAsSingles":False,"m_nTimeFormat":0,"m_nPrvyDen":0,"m_bPrintDayAsNumber":False,"m_DozoryKriteria":{"ucitel_limit_pocet":0,"ucitel_limit_minuty":0,"ucitel_rovnomerne_pocet":0,"ucitel_rovnomerne_minuty":0,"ucitel_uci_predalebopo":0,"ucitel_uci_predajpo":0,"ucitel_uci_vsusednej_predalebopo":0,"ucitel_uci_vsusednej_predajpo":0,"ucitel_cezdvojhodinovku":0,"ucitel_prazdnyden":0,"ucitel_predvyucovanim":0,"ucitel_povyucovani":0,"ucitel_vjedenden":0,"ucitel_zasebou":0},"draft_options":{}},"year":2022,"reg_name":""}],"data_columns":["name","settings","year","reg_name"]},
                        {"id":"periods","def":{"id":"periods","name":"Периоды","item_name":"Периоды","icon":"/global/pics/ui/time_32.svg"},"cdefs":[{"id":"period","type":"tableid","name":"Период","table":"periods"},{"id":"name","type":"string","name":"Название"},{"id":"short","type":"string","name":"Сокращение"},{"id":"starttime","type":"time","name":"Начинается"},{"id":"endtime","type":"time","name":"Заканчивается"},{"id":"daydata","type":"subdata","name":"daydata","subcolumns":[{"id":"starttime","type":"time","name":"Начинается"},{"id":"endtime","type":"time","name":"Заканчивается"}]},{"id":"printinsummary","type":"checkbox","name":"Печатать этот период в общих расписаниях"},{"id":"printinteacher","type":"checkbox","name":"Печатать этот период в отд. расп. учителей"},{"id":"printinclass","type":"checkbox","name":"Печатать этот период в отд. расп. учеников"},{"id":"printinclassroom","type":"checkbox","name":"Печатать этот период в отд. расп. кабинетов"},{"id":"printonlyinbells","type":"tableids","name":"Печат. в звонках","table":"bells"}],"data_rows":[
                            {
                                "id": str(i),
                                "period": period["period"],
                                "name": period["name"],
                                "short": period["short"],
                                "starttime": period["starttime"],
                                "endtime": period["endtime"],
                                "daydata": {},
                                "printinsummary": True,
                                "printinteacher": True,
                                "printinclass": True,
                                "printinclassroom": True,
                                "printonlyinbells": []
                            }
                            for i, period in enumerate(self._periods, 1)
                        ], "data_columns":["period","name","short","starttime","endtime","daydata","printinsummary","printinteacher","printinclass","printinclassroom","printonlyinbells"]},
                        {"id":"breaks","def":{"id":"breaks","name":"Перерывы","item_name":"Перерыв","icon":"/global/pics/ui/break_32.png"},"cdefs":[{"id":"break","type":"tableid","name":"перемена","table":"breaks"},{"id":"name","type":"string","name":"Название"},{"id":"short","type":"string","name":"Сокращение"},{"id":"starttime","type":"time","name":"Начинается"},{"id":"endtime","type":"time","name":"Заканчивается"},{"id":"daydata","type":"subdata","name":"daydata","subcolumns":[{"id":"starttime","type":"time","name":"Начинается"},{"id":"endtime","type":"time","name":"Заканчивается"}]},{"id":"printinsummary","type":"checkbox","name":"Печатать этот период в общих расписаниях"},{"id":"printinteacher","type":"checkbox","name":"Печатать этот период в отд. расп. учителей"},{"id":"printinclass","type":"checkbox","name":"Печатать этот период в отд. расп. учеников"},{"id":"printinclassroom","type":"checkbox","name":"Печатать этот период в отд. расп. кабинетов"},{"id":"printonlyinbells","type":"tableids","name":"Печат. в звонках","table":"bells"},{"id":"printtext","type":"string","name":"Текст для распечатки"}],"data_rows":[],"data_columns":["break","name","short","starttime","endtime","daydata","printinsummary","printinteacher","printinclass","printinclassroom","printonlyinbells","printtext"]},
                        {"id":"bells","def":{"id":"bells","name":"bells","item_name":"bells","icon":""},"cdefs":[{"id":"bell","type":"tableid","name":"Звонки","table":"bells"},{"id":"perioddata","type":"subdata","name":"perioddata","subcolumns":[{"id":"starttime","type":"time","name":"Начинается"},{"id":"endtime","type":"time","name":"Заканчивается"},{"id":"daydata","type":"subdata","name":"daydata","subcolumns":[{"id":"starttime","type":"time","name":"Начинается"},{"id":"endtime","type":"time","name":"Заканчивается"}]}]},{"id":"breakdata","type":"subdata","name":"breakdata","subcolumns":[{"id":"starttime","type":"time","name":"Начинается"},{"id":"endtime","type":"time","name":"Заканчивается"},{"id":"daydata","type":"subdata","name":"daydata","subcolumns":[{"id":"starttime","type":"time","name":"Начинается"},{"id":"endtime","type":"time","name":"Заканчивается"}]}]}],"data_rows":[{"id":"0","bell":"0","perioddata":{},"breakdata":{}},{"id":"1","bell":"1","perioddata":{},"breakdata":{}},{"id":"2","bell":"2","perioddata":{},"breakdata":{}},{"id":"3","bell":"3","perioddata":{},"breakdata":{}},{"id":"4","bell":"4","perioddata":{},"breakdata":{}},{"id":"5","bell":"5","perioddata":{},"breakdata":{}},{"id":"6","bell":"6","perioddata":{},"breakdata":{}},{"id":"7","bell":"7","perioddata":{},"breakdata":{}},{"id":"8","bell":"8","perioddata":{},"breakdata":{}},{"id":"9","bell":"9","perioddata":{},"breakdata":{}},{"id":"10","bell":"10","perioddata":{},"breakdata":{}},{"id":"11","bell":"11","perioddata":{},"breakdata":{}},{"id":"12","bell":"12","perioddata":{},"breakdata":{}},{"id":"13","bell":"13","perioddata":{},"breakdata":{}},{"id":"14","bell":"14","perioddata":{},"breakdata":{}},{"id":"15","bell":"15","perioddata":{},"breakdata":{}},{"id":"16","bell":"16","perioddata":{},"breakdata":{}},{"id":"17","bell":"17","perioddata":{},"breakdata":{}},{"id":"18","bell":"18","perioddata":{},"breakdata":{}},{"id":"19","bell":"19","perioddata":{},"breakdata":{}},{"id":"20","bell":"20","perioddata":{},"breakdata":{}},{"id":"21","bell":"21","perioddata":{},"breakdata":{}},{"id":"22","bell":"22","perioddata":{},"breakdata":{}},{"id":"23","bell":"23","perioddata":{},"breakdata":{}},{"id":"24","bell":"24","perioddata":{},"breakdata":{}},{"id":"25","bell":"25","perioddata":{},"breakdata":{}},{"id":"26","bell":"26","perioddata":{},"breakdata":{}},{"id":"27","bell":"27","perioddata":{},"breakdata":{}},{"id":"28","bell":"28","perioddata":{},"breakdata":{}},{"id":"29","bell":"29","perioddata":{},"breakdata":{}},{"id":"30","bell":"30","perioddata":{},"breakdata":{}},{"id":"31","bell":"31","perioddata":{},"breakdata":{}},{"id":"32","bell":"32","perioddata":{},"breakdata":{}},{"id":"33","bell":"33","perioddata":{},"breakdata":{}},{"id":"34","bell":"34","perioddata":{},"breakdata":{}},{"id":"35","bell":"35","perioddata":{},"breakdata":{}},{"id":"36","bell":"36","perioddata":{},"breakdata":{}},{"id":"37","bell":"37","perioddata":{},"breakdata":{}},{"id":"38","bell":"38","perioddata":{},"breakdata":{}},{"id":"39","bell":"39","perioddata":{},"breakdata":{}},{"id":"40","bell":"40","perioddata":{},"breakdata":{}},{"id":"41","bell":"41","perioddata":{},"breakdata":{}},{"id":"42","bell":"42","perioddata":{},"breakdata":{}},{"id":"43","bell":"43","perioddata":{},"breakdata":{}},{"id":"44","bell":"44","perioddata":{},"breakdata":{}},{"id":"45","bell":"45","perioddata":{},"breakdata":{}},{"id":"46","bell":"46","perioddata":{},"breakdata":{}},{"id":"47","bell":"47","perioddata":{},"breakdata":{}},{"id":"48","bell":"48","perioddata":{},"breakdata":{}},{"id":"49","bell":"49","perioddata":{},"breakdata":{}}],"data_columns":["bell","perioddata","breakdata"]},
                        {"id":"daysdefs","def":{"id":"daysdefs","name":"Дни","item_name":"День","icon":"/timetable/pics/app/term_week_day/r_2_A.png"},"cdefs":[{"id":"name","type":"string","name":"Название"},{"id":"short","type":"string","name":"Сокращение"},{"id":"typ","type":"enum","name":"typ"},{"id":"vals","type":"stringarray","name":"vals"},{"id":"val","type":"int","name":"val"}],"data_rows":[
                            {
                                "id": "*{}".format(i),
                                "name": daysdef["name"],
                                "short": daysdef["short"],
                                "typ": daysdef["typ"],
                                "vals": daysdef["days"].split(","),
                                "val": daysdef["val"]
                            }
                            for i, daysdef in enumerate(self._daysdefs, 1)
                        ],"data_columns":["name","short","typ","vals","val"]},
                        {"id":"weeksdefs","def":{"id":"weeksdefs","name":"Недели","item_name":"Неделя","icon":"/timetable/pics/app/term_week_day/g_2_A.png"},"cdefs":[{"id":"name","type":"string","name":"Название"},{"id":"short","type":"string","name":"Сокращение"},{"id":"typ","type":"enum","name":"typ"},{"id":"vals","type":"stringarray","name":"vals"},{"id":"val","type":"int","name":"val"}],"data_rows":[{"id":"*1","name":"Неделя A","short":"A","typ":"one","vals":["1"],"val":0},{"id":"*2","name":"Любая неделя","short":"Любой","typ":"any","vals":["1"],"val":None},{"id":"*3","name":"Все недели","short":"Все","typ":"all","vals":["1"],"val":None}],"data_columns":["name","short","typ","vals","val"]},
                        {"id":"termsdefs","def":{"id":"termsdefs","name":"Семестры","item_name":"Семестр","icon":"/timetable/pics/app/term_week_day/b_2_A.png"},"cdefs":[{"id":"name","type":"string","name":"Название"},{"id":"short","type":"string","name":"Сокращение"},{"id":"typ","type":"enum","name":"typ"},{"id":"vals","type":"stringarray","name":"vals"},{"id":"val","type":"int","name":"val"}],"data_rows":[{"id":"*1","name":"Весь год","short":"Год","typ":"all","vals":["1"],"val":None}],"data_columns":["name","short","typ","vals","val"]},
                        {"id":"days","def":{"id":"days","name":"Дни","item_name":"День","icon":"/timetable/pics/app/term_week_day/r_2_A.png"},"cdefs":[{"id":"name","type":"string","name":"Название"},{"id":"short","type":"string","name":"Сокращение"}],"data_rows":[
                            {
                                "id": str(i),
                                "name": daysdef["name"],
                                "short": daysdef["short"]
                            }
                            for i, daysdef in enumerate(self._daysdefs, 0)
                        ],"data_columns":["name","short"]},
                        {"id":"weeks","def":{"id":"weeks","name":"Недели","item_name":"Неделя","icon":"/timetable/pics/app/term_week_day/g_2_A.png"},"cdefs":[{"id":"name","type":"string","name":"Название"},{"id":"short","type":"string","name":"Сокращение"}],"data_rows":[{"id":"0","name":"Неделя A","short":"A"}],"data_columns":["name","short"]},
                        {"id":"terms","def":{"id":"terms","name":"Семестры","item_name":"Семестр","icon":"/timetable/pics/app/term_week_day/b_2_A.png"},"cdefs":[{"id":"name","type":"string","name":"Название"},{"id":"short","type":"string","name":"Сокращение"}],"data_rows":[{"id":"0","name":"Семестр 1","short":"T1"}],"data_columns":["name","short"]},
                        {"id":"buildings","def":{"id":"buildings","name":"Школьные корпуса","item_name":"Учебный корпус","icon":"/global/pics/ui/classroom_32.svg"},"cdefs":[{"id":"name","type":"string","name":"Название"},{"id":"short","type":"string","name":"Сокращение"},{"id":"color","type":"color","name":"Цвет"}],"data_rows":[],"data_columns":["name","short","color"]},
                        {"id":"classrooms","def":{"id":"classrooms","name":"Кабинеты","item_name":"Кабинет","icon":"/global/pics/ui/classroom_32.svg"},"cdefs":[{"id":"name","type":"string","name":"Название"},{"id":"short","type":"string","name":"Сокращение"},{"id":"buildingid","type":"tableid","name":"Учебный корпус","table":"buildings"},{"id":"sharedroom","type":"checkbox","name":"Общий кабинет"},{"id":"needssupervision","type":"checkbox","name":"Этому кабинету нужен надсмотр"},{"id":"color","type":"color","name":"Цвет"},{"id":"nearbyclassroomids","type":"tableids","name":"Кабинеты поблизости","table":"classrooms"}],"data_rows":[
                            classroom.dict()
                            for classroom in self._classrooms
                        ],"data_columns":["name","short","buildingid","sharedroom","needssupervision","color","nearbyclassroomids"]},
                        {"id":"classes","def":{"id":"classes","name":"Классы","item_name":"Класс","icon":"/global/pics/ui/class_32.svg"},"cdefs":[{"id":"name","type":"string","name":"Название"},{"id":"short","type":"string","name":"Сокращение"},{"id":"teacherid","type":"tableid","name":"Классный руководитель","table":"teachers"},{"id":"classroomids","type":"tableids","name":"Основной кабинет","table":"classrooms"},{"id":"bell","type":"tableid","name":"Звонки","table":"bells"},{"id":"color","type":"color","name":"Цвет"},{"id":"timeoff","type":"object","name":"Рабочее время"},{"id":"printsubjectpictures","type":"checkbox","name":"Печатать картинки предметов"},{"id":"classroomid","type":"tableid","name":"Кабинет","table":"classrooms"}],"data_rows":[
                            class_.dict()
                            for class_ in self._classes
                        ],"data_columns":["name","short","teacherid","classroomids","bell","color","timeoff","printsubjectpictures","classroomid"]},
                        {"id":"subjects","def":{"id":"subjects","name":"Предметы","item_name":"Предмет","icon":"/global/pics/ui/subject_32.svg"},"cdefs":[{"id":"name","type":"string","name":"Название"},{"id":"short","type":"string","name":"Сокращение"},{"id":"color","type":"color","name":"Цвет"},{"id":"picture_url","name":"Картинка"},{"id":"timeoff","type":"object","name":"Рабочее время"},{"id":"contract_weight","type":"float","name":"Продолжительность контракта учителя"}],"data_rows":[
                            subject.dict()
                            for subject in self._subjects
                        ],"data_columns":["name","short","color","picture_url","timeoff","contract_weight"]},
                        {"id":"teachers","def":{"id":"teachers","name":"Учителя","item_name":"Учитель","icon":"/global/pics/ui/teacher_32.svg"},"cdefs":[{"id":"firstname","type":"string","name":"Имя"},{"id":"lastname","type":"string","name":"Фамилия"},{"id":"nameprefix","type":"string","name":"Обращение"},{"id":"namesuffix","type":"string","name":"суффикс имени"},{"id":"short","type":"string","name":"Сокращение"},{"id":"gender","type":"enum","name":"Пол"},{"id":"bell","type":"tableid","name":"Звонки","table":"bells"},{"id":"color","type":"color","name":"Цвет"},{"id":"fontcolorprint","type":"color","name":"Основной текст"},{"id":"fontcolorprint2","type":"color","name":"Другой текст"},{"id":"fontcolorscreen","type":"color","name":"Текст на экране"},{"id":"timeoff","type":"object","name":"Рабочее время"}],"data_rows":[
                            teacher.dict()
                            for teacher in self._teachers
                        ],"data_columns":["firstname","lastname","nameprefix","namesuffix","short","gender","bell","color","fontcolorprint","fontcolorprint2","fontcolorscreen","timeoff"]},
                        {"id":"groups","def":{"id":"groups","name":"Группы","item_name":"Группа","icon":"/global/pics/ui/group_32.svg"},"cdefs":[{"id":"name","type":"string","name":"Название"},{"id":"classid","type":"tableid","name":"Класс","table":"classes"},{"id":"entireclass","type":"checkbox","name":"entireclass"},{"id":"ascttdivision","type":"string","name":"ascttdivision"},{"id":"okgroup","name":"okgroup"},{"id":"students_count","type":"int","name":"students_count"},{"id":"divisionid","type":"tableid","name":"divisionid","table":"divisions"},{"id":"color","type":"color","name":"Цвет"}],"data_rows":[
                            group.dict()
                            for group in self._groups
                        ],"data_columns":["name","classid","entireclass","ascttdivision","okgroup","students_count","divisionid","color"]},
                        {"id":"divisions","def":{"id":"divisions","name":"divisions","item_name":"divisions","icon":"/timetable/pics/app/office/division_32.svg"},"cdefs":[{"id":"groupids","type":"tableids","name":"Группы","table":"groups"}],"data_rows":list(itertools_chain(*
                            [
                                [
                                    {
                                        "id": "*{}:".format(i),
                                        "groupids": [
                                            groups[0].id
                                        ]
                                    },
                                    {
                                        "id": "*{}:1".format(i),
                                        "groupids": [
                                            group.id
                                            for group in groups[1:]
                                        ]
                                    }
                                ]
                                for i, groups in enumerate(self._groups_from_class_id.values(), 1)
                            ]
                        )),"data_columns":["groupids"]},
                        {"id":"students","def":{"id":"students","name":"Учащиеся","item_name":"Учащийся","icon":"/global/pics/ui/student_32.svg"},"cdefs":[{"id":"classid","type":"tableid","name":"Класс","table":"classes"},{"id":"groupids","type":"tableids","name":"Группы","table":"groups"},{"id":"short","type":"string","name":"Сокращение"}],"data_rows":[],"data_columns":["classid","groupids","short"]},
                        {"id":"lessons","def":{"id":"lessons","name":"Уроки","item_name":"Урок","icon":"/global/pics/ui/lessons_32.svg"},"cdefs":[{"id":"subjectid","type":"tableid","name":"Предмет","table":"subjects"},{"id":"teacherids","type":"tableids","name":"Учителя","table":"teachers"},{"id":"groupids","type":"tableids","name":"Группы","table":"groups"},{"id":"classids","type":"tableids","name":"Класс","table":"classes"},{"id":"count","type":"int","name":"Всего"},{"id":"durationperiods","type":"int","name":"Длина"},{"id":"classroomidss","type":"tableidss","name":"Свободные кабинеты","table":"classrooms"},{"id":"termsdefid","type":"tableid","name":"Семестр","table":"termsdefs"},{"id":"weeksdefid","type":"tableid","name":"Неделя","table":"weeksdefs"},{"id":"daysdefid","type":"tableid","name":"День недели","table":"daysdefs"},{"id":"terms","name":"terms"},{"id":"seminargroup","type":"int","name":"Группа семинара №"},{"id":"bell","type":"tableid","name":"Звонки","table":"bells"},{"id":"studentids","type":"tableids","name":"Учащиеся","table":"students"},{"id":"groupnames","type":"stringarray","name":"Группы"}],"data_rows":[
                            lesson.dict()
                            for lesson in self._lessons
                        ],"data_columns":["subjectid","teacherids","groupids","classids","count","durationperiods","classroomidss","termsdefid","weeksdefid","daysdefid","terms","seminargroup","bell","studentids","groupnames"]},
                        {"id":"studentsubjects","def":{"id":"studentsubjects","name":"Семинары","item_name":"Семинары","icon":"/timetable/pics/app/office/seminar_32.png"},"cdefs":[{"id":"studentid","type":"tableid","name":"Учащийся","table":"students"},{"id":"subjectid","type":"tableid","name":"Предмет","table":"subjects"},{"id":"seminargroup","type":"int","name":"Группа семинара №"},{"id":"importance","type":"enum","name":"Важность"},{"id":"locked","type":"checkbox","name":"Запертый"}],"data_rows":[],"data_columns":["studentid","subjectid","seminargroup","importance","locked"]},
                        {"id":"cards","def":{"id":"cards","name":"Карточки","item_name":"Карточка урока","icon":"/timetable/pics/app/card_32.png"},"cdefs":[{"id":"lessonid","type":"tableid","name":"lessonid","table":"lessons"},{"id":"period","type":"tableid","name":"Период","table":"periods"},{"id":"days","type":"string","name":"days"},{"id":"weeks","type":"string","name":"weeks"},{"id":"classroomids","type":"tableids","name":"Кабинеты","table":"classrooms"}],"data_rows":[
                            card.dict()
                            for card in self._cards
                        ],"data_columns":["lessonid","period","days","weeks","classroomids"]},
                        {"id":"ttreports","def":{"id":"ttreports","name":"Сводки","item_name":"Сводки","icon":"/timetable/pics/app/print_preview_32.png"},"cdefs":[{"id":"typ","type":"int","name":"typ"},{"id":"fitwidth","type":"checkbox","name":"Устан. ширину в одну страницу"},{"id":"fitheight","type":"checkbox","name":"Устан. высоту в одну страницу"},{"id":"hideemptycolumns","type":"checkbox","name":"Скрыть пустые столбцы"},{"id":"hideemptyrows","type":"checkbox","name":"Скрыть пустые строки"},{"id":"headerwidth","type":"float","name":"headerwidth"},{"id":"headerheight","type":"float","name":"headerheight"},{"id":"cellwidth","type":"float","name":"cellwidth"},{"id":"cellheight","type":"float","name":"cellheight"},{"id":"page_tables","type":"enumarray","name":"Печатать одну страницу для"},{"id":"row_tables","type":"enumarray","name":"Строки"},{"id":"column_tables","type":"enumarray","name":"Столбцы"},{"id":"celltype","type":"intcombo","name":"Ячейки"},{"id":"cardcolorenabled","type":"checkbox","name":"Цветная печать"},{"id":"cardcolortable1","type":"enum","name":"cardcolortable1"},{"id":"cardcolortable2","type":"enum","name":"cardcolortable2"},{"id":"cardcolorpos","name":"cardcolorpos"},{"id":"cardstyles","type":"subarray","name":"Карточки - Стиль","subcolumns":[{"id":"m_nDlzka","type":"int","name":"m_nDlzka"},{"id":"m_nPocetRiadkov","type":"int","name":"m_nPocetRiadkov"},{"id":"m_nBezTriedyAleboUcitela","type":"int","name":"m_nBezTriedyAleboUcitela"},{"id":"texts","type":"subdata","name":"texts","subcolumns":[{"id":"enabled","type":"checkbox","name":"enabled"},{"id":"pos","type":"int","name":"pos"},{"id":"size","type":"float","name":"Размер"},{"id":"font","type":"string","name":"Шрифт"},{"id":"bold","type":"checkbox","name":"Жирный шрифт"},{"id":"italic","type":"checkbox","name":"Курсив"},{"id":"underline","type":"checkbox","name":"Подчеркнутый шрифт"},{"id":"autohide","type":"checkbox","name":"autohide"},{"id":"name_col","type":"string","name":"name_col"}]}]},{"id":"classroomsupervisionstyle","type":"subobject","name":"classroomsupervisionstyle","subcolumns":[{"id":"size","type":"float","name":"Размер"},{"id":"font","type":"string","name":"Шрифт"},{"id":"bold","type":"checkbox","name":"Жирный шрифт"},{"id":"italic","type":"checkbox","name":"Курсив"},{"id":"underline","type":"checkbox","name":"Подчеркнутый шрифт"}]},{"id":"gridheadertexts","type":"subdata","name":"Название - Стиль","subcolumns":[{"id":"enabled","type":"boolean_or_null","name":"enabled"},{"id":"pos","type":"int","name":"pos"},{"id":"size","type":"float","name":"Размер"},{"id":"font","type":"string","name":"Шрифт"},{"id":"bold","type":"checkbox","name":"Жирный шрифт"},{"id":"italic","type":"checkbox","name":"Курсив"},{"id":"underline","type":"checkbox","name":"Подчеркнутый шрифт"},{"id":"column","type":"string","name":"column"}]},{"id":"pageheader","type":"subobject","name":"Название","subcolumns":[{"id":"font","type":"string","name":"Шрифт"},{"id":"size","type":"float","name":"Размер"},{"id":"bold","type":"checkbox","name":"Жирный шрифт"},{"id":"italic","type":"checkbox","name":"Курсив"},{"id":"underline","type":"checkbox","name":"Подчеркнутый шрифт"}]},{"id":"pageheaderprefixenabled","type":"checkbox","name":"Печатать префикс"},{"id":"userheader","type":"subobject","name":"userheader","subcolumns":[{"id":"text","type":"string","name":"text"},{"id":"font","type":"string","name":"Шрифт"},{"id":"size","type":"float","name":"Размер"},{"id":"bold","type":"checkbox","name":"Жирный шрифт"},{"id":"italic","type":"checkbox","name":"Курсив"},{"id":"underline","type":"checkbox","name":"Подчеркнутый шрифт"}]},{"id":"settings","name":"settings"},{"id":"landscape","type":"checkbox","name":"landscape"},{"id":"repeatpage","type":"int","name":"repeatpage"},{"id":"splitpage_h","type":"int","name":"splitpage_h"},{"id":"splitpage_w","type":"int","name":"splitpage_w"},{"id":"withclassroomtt","type":"checkbox","name":"withclassroomtt"},{"id":"printlogo","type":"checkbox","name":"printlogo"},{"id":"name","type":"string","name":"Название"},{"id":"extracolumns","type":"subarray","name":"Дополнительные столбцы","subcolumns":[{"id":"typ","type":"combo","name":"typ"},{"id":"name","type":"string","name":"Название"},{"id":"size","type":"int","name":"Размер"},{"id":"headerstyle","type":"subobject","name":"headerstyle","subcolumns":[{"id":"size","type":"float","name":"Размер"},{"id":"font","type":"string","name":"Шрифт"},{"id":"bold","type":"checkbox","name":"Жирный шрифт"},{"id":"italic","type":"checkbox","name":"Курсив"},{"id":"underline","type":"checkbox","name":"Подчеркнутый шрифт"},{"id":"vertical","type":"checkbox","name":"vertical"}]},{"id":"cellstyle","type":"subobject","name":"cellstyle","subcolumns":[{"id":"size","type":"float","name":"Размер"},{"id":"font","type":"string","name":"Шрифт"},{"id":"bold","type":"checkbox","name":"Жирный шрифт"},{"id":"italic","type":"checkbox","name":"Курсив"},{"id":"underline","type":"checkbox","name":"Подчеркнутый шрифт"},{"id":"vertical","type":"checkbox","name":"vertical"}]}]},{"id":"extrarows","type":"subarray","name":"Дополнительные строки","subcolumns":[{"id":"typ","type":"combo","name":"typ"},{"id":"name","type":"string","name":"Название"},{"id":"size","type":"int","name":"Размер"},{"id":"headerstyle","type":"subobject","name":"headerstyle","subcolumns":[{"id":"size","type":"float","name":"Размер"},{"id":"font","type":"string","name":"Шрифт"},{"id":"bold","type":"checkbox","name":"Жирный шрифт"},{"id":"italic","type":"checkbox","name":"Курсив"},{"id":"underline","type":"checkbox","name":"Подчеркнутый шрифт"},{"id":"vertical","type":"checkbox","name":"vertical"}]},{"id":"cellstyle","type":"subobject","name":"cellstyle","subcolumns":[{"id":"size","type":"float","name":"Размер"},{"id":"font","type":"string","name":"Шрифт"},{"id":"bold","type":"checkbox","name":"Жирный шрифт"},{"id":"italic","type":"checkbox","name":"Курсив"},{"id":"underline","type":"checkbox","name":"Подчеркнутый шрифт"},{"id":"vertical","type":"checkbox","name":"vertical"}]}]}],"data_rows":[],"data_columns":["typ","fitwidth","fitheight","hideemptycolumns","hideemptyrows","headerwidth","headerheight","cellwidth","cellheight","page_tables","row_tables","column_tables","celltype","cardcolorenabled","cardcolortable1","cardcolortable2","cardcolorpos","cardstyles","classroomsupervisionstyle","gridheadertexts","pageheader","pageheaderprefixenabled","userheader","settings","landscape","repeatpage","splitpage_h","splitpage_w","withclassroomtt","printlogo","name","extracolumns","extrarows"]},
                        {"id":"classroomsupervisions","def":{"id":"classroomsupervisions","name":"Надсмотры","item_name":"Надсмотр","icon":"/timetable/pics/app/views/supervisions.png"},"cdefs":[{"id":"classroomid","type":"tableid","name":"Кабинет","table":"classrooms"},{"id":"teacherid","type":"tableid","name":"Учитель","table":"teachers"},{"id":"day","name":"day"},{"id":"week","name":"week"},{"id":"term","name":"term"},{"id":"period","type":"tableid","name":"Период","table":"periods"},{"id":"break","type":"tableid","name":"перемена","table":"breaks"},{"id":"locked","type":"checkbox","name":"Запертый"},{"id":"weeks","name":"weeks"}],"data_rows":[],"data_columns":["classroomid","teacherid","day","week","term","period","break","locked","weeks"]},
                        {"id":"coursegroups","def":{"id":"coursegroups","name":"Группы семинаров","item_name":"Группа семинара","icon":"/timetable/pics/app/office/sectiongroups_32.svg"},"cdefs":[{"id":"name","type":"string","name":"Название"}],"data_rows":[],"data_columns":["name"]}
                    ]
                }
            }
        }

        with open(json_filepath, "w", encoding="utf-8") as file:
            dump(
                obj = timetable_dict,
                fp = file,
                ensure_ascii = False,
                indent = json_indent
            )
